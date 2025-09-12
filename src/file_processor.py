"""
File processing module for 90D-PPT Generator
Handles Excel and CSV file reading, sheet detection, and data processing
"""

import pandas as pd
import streamlit as st
from typing import Dict, List, Tuple, Optional, Any
from pathlib import Path
import io
import openpyxl
from openpyxl import load_workbook

from utils import (
    detect_header_row, normalize_column_name, categorize_status,
    normalize_ranking, show_error_message, show_warning_message, show_success_message
)


class FileProcessor:
    """Handles file processing operations for Excel and CSV files"""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls', '.csv']
        self.max_file_size_mb = 200
    
    def read_file(self, uploaded_file) -> Tuple[Optional[pd.DataFrame], List[str], str]:
        """
        Read uploaded file and return DataFrame, sheet names, and any error message
        
        Returns:
            tuple: (DataFrame, list of sheet names, error message)
        """
        try:
            file_extension = Path(uploaded_file.name).suffix.lower()
            
            if file_extension == '.csv':
                return self._read_csv(uploaded_file)
            elif file_extension in ['.xlsx', '.xls']:
                return self._read_excel(uploaded_file)
            else:
                return None, [], f"Unsupported file type: {file_extension}"
                
        except Exception as e:
            return None, [], f"Error reading file: {str(e)}"
    
    def _read_csv(self, uploaded_file) -> Tuple[Optional[pd.DataFrame], List[str], str]:
        """Read CSV file"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            df = None
            
            for encoding in encodings:
                try:
                    uploaded_file.seek(0)
                    df = pd.read_csv(
                        uploaded_file, 
                        encoding=encoding,
                        dtype=str,  # Read all as strings initially
                        keep_default_na=False
                    )
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                return None, [], "Could not decode CSV file with any supported encoding"
            
            # Clean up the dataframe
            df = self._clean_dataframe(df)
            
            return df, ['Sheet1'], ""  # CSV files only have one "sheet"
            
        except Exception as e:
            return None, [], f"Error reading CSV: {str(e)}"
    
    def _read_excel(self, uploaded_file) -> Tuple[Optional[pd.DataFrame], List[str], str]:
        """Enhanced Excel file reading with better handling of complex structures"""
        try:
            # First, get sheet names and metadata using openpyxl
            uploaded_file.seek(0)
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            
            if not sheet_names:
                workbook.close()
                return None, [], "No sheets found in Excel file"
            
            # Get the first sheet for inspection
            first_sheet = workbook[sheet_names[0]]
            workbook.close()
            
            # Read with multiple strategies to handle complex Excel files
            uploaded_file.seek(0)
            
            # Try reading with header=None first to preserve structure
            df = pd.read_excel(
                uploaded_file, 
                sheet_name=sheet_names[0],
                header=None,  # Don't assume first row is header
                dtype=str,
                keep_default_na=False
            )
            
            if df.empty:
                return None, [], "Excel sheet is empty"
            
            # Clean up the dataframe with enhanced logic
            df = self._clean_dataframe_enhanced(df)
            
            return df, sheet_names, ""
            
        except Exception as e:
            return None, [], f"Error reading Excel: {str(e)}"
    
    def read_excel_sheet(self, uploaded_file, sheet_name: str) -> Optional[pd.DataFrame]:
        """Read specific sheet from Excel file with enhanced processing"""
        try:
            uploaded_file.seek(0)
            df = pd.read_excel(
                uploaded_file, 
                sheet_name=sheet_name,
                header=None,  # Don't assume first row is header
                dtype=str,
                keep_default_na=False
            )
            
            return self._clean_dataframe_enhanced(df)
            
        except Exception as e:
            show_error_message(f"Error reading sheet '{sheet_name}': {str(e)}")
            return None
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and prepare dataframe for processing (legacy method)"""
        if df.empty:
            return df
        
        # Remove completely empty rows and columns
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # Clean column names - handle unnamed columns properly
        clean_columns = []
        for i, col in enumerate(df.columns):
            if pd.notna(col) and str(col).strip():
                # Clean valid column names
                clean_col = str(col).strip()
                # Handle "Unnamed: X" patterns from pandas
                if clean_col.startswith('Unnamed:'):
                    clean_col = f"Column_{i+1}"
                clean_columns.append(clean_col)
            else:
                # Handle empty or None column names
                clean_columns.append(f"Column_{i+1}")
        
        df.columns = clean_columns
        
        # Remove duplicate column names by adding suffix
        df.columns = pd.io.common.dedup_names(df.columns, is_potential_multiindex=False)
        
        # Convert empty strings to None for better handling
        df = df.replace('', None)
        
        return df
        
    def _clean_dataframe_enhanced(self, df: pd.DataFrame) -> pd.DataFrame:
        """Enhanced dataframe cleaning with better Excel handling"""
        if df.empty:
            return df
        
        # Find the actual data boundaries by scanning for non-empty content
        first_data_row = 0
        last_data_row = len(df) - 1
        first_data_col = 0
        last_data_col = len(df.columns) - 1
        
        # Find first row with meaningful data (not just whitespace)
        for i in range(len(df)):
            row_values = [str(val).strip() for val in df.iloc[i].values if pd.notna(val)]
            if any(val for val in row_values if val):  # Has non-empty values
                first_data_row = i
                break
        
        # Find last row with data
        for i in range(len(df) - 1, -1, -1):
            row_values = [str(val).strip() for val in df.iloc[i].values if pd.notna(val)]
            if any(val for val in row_values if val):
                last_data_row = i
                break
        
        # Find first and last columns with data
        for j in range(len(df.columns)):
            col_values = [str(val).strip() for val in df.iloc[:, j].values if pd.notna(val)]
            if any(val for val in col_values if val):
                first_data_col = j
                break
                
        for j in range(len(df.columns) - 1, -1, -1):
            col_values = [str(val).strip() for val in df.iloc[:, j].values if pd.notna(val)]
            if any(val for val in col_values if val):
                last_data_col = j
                break
        
        # Crop to data boundaries
        df = df.iloc[first_data_row:last_data_row+1, first_data_col:last_data_col+1].copy()
        
        # Reset index
        df = df.reset_index(drop=True)
        
        # Temporarily set column names as strings for processing
        df.columns = [f"Column_{i}" for i in range(len(df.columns))]
        
        # Convert empty strings and whitespace to None
        df = df.map(lambda x: None if pd.isna(x) or str(x).strip() == '' else str(x).strip())
        
        return df
    
    def detect_headers(self, df: pd.DataFrame) -> int:
        """Detect header row in dataframe"""
        return detect_header_row(df)
    
    def apply_header_row(self, df: pd.DataFrame, header_row: int) -> pd.DataFrame:
        """Enhanced header row application with better name handling"""
        if df.empty:
            return df
            
        try:
            if header_row >= len(df):
                header_row = 0  # Fallback to first row
            
            if header_row > 0:
                # Extract header row
                header_values = df.iloc[header_row].values
                
                # Create new dataframe with data after header
                data_rows = df.iloc[header_row + 1:].copy()
                
                # Apply headers to columns
                data_rows.columns = range(len(data_rows.columns))  # Reset to numeric first
                
                # Create meaningful column names from header row
                clean_headers = []
                for i, header_val in enumerate(header_values):
                    if pd.notna(header_val) and str(header_val).strip():
                        # Clean and use the header value
                        clean_header = str(header_val).strip()
                        # Remove common Excel artifacts
                        clean_header = clean_header.replace('\n', ' ').replace('\r', ' ')
                        clean_header = ' '.join(clean_header.split())  # Normalize whitespace
                        clean_headers.append(clean_header)
                    else:
                        # Generate fallback name
                        clean_headers.append(f"Column_{i+1}")
                
                # Ensure no duplicate column names
                seen = set()
                final_headers = []
                for header in clean_headers:
                    original_header = header
                    counter = 1
                    while header in seen:
                        header = f"{original_header}_{counter}"
                        counter += 1
                    seen.add(header)
                    final_headers.append(header)
                
                # Apply the cleaned headers
                data_rows.columns = final_headers
                new_df = data_rows.reset_index(drop=True)
            else:
                # Use first row as header (header_row = 0)
                new_df = df.copy()
                # Generate column names from first row or use generic names
                if len(df) > 0:
                    header_values = df.iloc[0].values
                    new_df = df.iloc[1:].copy()
                    
                    clean_headers = []
                    for i, header_val in enumerate(header_values):
                        if pd.notna(header_val) and str(header_val).strip():
                            clean_header = str(header_val).strip()
                            clean_header = clean_header.replace('\n', ' ').replace('\r', ' ')
                            clean_header = ' '.join(clean_header.split())
                            clean_headers.append(clean_header)
                        else:
                            clean_headers.append(f"Column_{i+1}")
                    
                    # Handle duplicates
                    seen = set()
                    final_headers = []
                    for header in clean_headers:
                        original_header = header
                        counter = 1
                        while header in seen:
                            header = f"{original_header}_{counter}"
                            counter += 1
                        seen.add(header)
                        final_headers.append(header)
                    
                    new_df.columns = final_headers
                    new_df = new_df.reset_index(drop=True)
                else:
                    # Generate generic column names
                    new_df.columns = [f"Column_{i+1}" for i in range(len(new_df.columns))]
            
            return new_df
            
        except Exception as e:
            # Fallback: return dataframe with generic column names
            df.columns = [f"Column_{i+1}" for i in range(len(df.columns))]
            return df
    
    def get_data_preview(self, df: pd.DataFrame, max_rows: int = 100) -> pd.DataFrame:
        """Get preview of data for display"""
        if df.empty:
            return df
        
        # Return first max_rows for preview
        preview_df = df.head(max_rows)
        
        # Show data types and non-null counts
        return preview_df
    
    def get_data_summary(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Get summary statistics about the dataframe"""
        if df.empty:
            return {
                'total_rows': 0,
                'total_columns': 0,
                'columns': [],
                'data_types': {},
                'null_counts': {}
            }
        
        summary = {
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'columns': list(df.columns),
            'data_types': {col: str(df[col].dtype) for col in df.columns},
            'null_counts': df.isnull().sum().to_dict(),
            'non_null_counts': df.count().to_dict()
        }
        
        return summary
    
    def filter_by_rankings(self, df: pd.DataFrame, ranking_column: str, 
                          selected_rankings: List[str]) -> pd.DataFrame:
        """Filter dataframe by selected ranking values"""
        if df.empty or ranking_column not in df.columns:
            return df
        
        if not selected_rankings:
            return df
        
        # Normalize selected rankings for comparison
        normalized_selected = [r.lower() for r in selected_rankings]
        
        # Filter rows where ranking matches any of the selected values
        mask = df[ranking_column].apply(
            lambda x: normalize_ranking(x) is not None and 
                     normalize_ranking(x).lower() in normalized_selected
        )
        
        filtered_df = df[mask].copy()
        
        # Add normalized ranking column for easier processing
        filtered_df['_normalized_ranking'] = filtered_df[ranking_column].apply(normalize_ranking)
        
        return filtered_df
    
    def get_ranking_counts(self, df: pd.DataFrame, ranking_column: str) -> Dict[str, int]:
        """Get counts of each ranking value"""
        if df.empty or ranking_column not in df.columns:
            return {}
        
        # Get normalized rankings and count them
        normalized_rankings = df[ranking_column].apply(normalize_ranking)
        valid_rankings = normalized_rankings[normalized_rankings.notna()]
        
        return valid_rankings.value_counts().to_dict()
    
    def get_status_counts(self, df: pd.DataFrame, status_column: str) -> Dict[str, int]:
        """Get status counts categorized as closed/non-closed"""
        if df.empty or status_column not in df.columns:
            return {'closed': 0, 'non-closed': 0}
        
        status_categories = df[status_column].apply(categorize_status)
        return status_categories.value_counts().to_dict()
    
    def prepare_data_for_ppt(self, df: pd.DataFrame, column_mapping: Dict[str, str]) -> Dict[str, pd.DataFrame]:
        """
        Prepare data for PowerPoint generation by ranking
        
        Returns:
            Dict with ranking as key and DataFrame as value
        """
        if df.empty:
            return {}
        
        # Get the ranking column - now optional
        ranking_col = column_mapping.get('ranking')
        result = {}
        
        if ranking_col and ranking_col in df.columns:
            # Group by normalized ranking in PRD-specified order
            ranking_order = ['Accepted', 'Up Next', 'Maybe', 'Likely No']
            
            for ranking in ranking_order:
                # Filter data for this ranking
                ranking_data = df[
                    df[ranking_col].apply(
                        lambda x: normalize_ranking(x) == ranking if pd.notna(x) else False
                    )
                ].copy()
                
                if not ranking_data.empty:
                    # Create PowerPoint data structure
                    ppt_data = self._create_ppt_dataframe(ranking_data, column_mapping)
                    result[ranking] = ppt_data
        else:
            # No ranking column mapped - create a single slide with "General" ranking
            # This allows users to generate presentations without ranking data
            general_data = df.copy()
            if not general_data.empty:
                ppt_data = self._create_ppt_dataframe(general_data, column_mapping)
                result['General'] = ppt_data
        
        return result
    
    def _create_ppt_dataframe(self, data: pd.DataFrame, column_mapping: Dict[str, str]) -> pd.DataFrame:
        """Create PowerPoint-ready DataFrame from source data and column mapping"""
        ppt_data = pd.DataFrame()
        
        # Map all columns according to mapping (all are optional)
        column_map = {
            'jira': column_mapping.get('jira', ''),
            'description': column_mapping.get('description', ''),
            'status': column_mapping.get('status', ''),
            'component': column_mapping.get('component', ''),
            'risks': column_mapping.get('risks', '')
        }
        
        for target_col, source_col in column_map.items():
            if source_col and source_col in data.columns:
                ppt_data[target_col] = data[source_col].fillna('')
            else:
                ppt_data[target_col] = ""  # Empty column if not mapped
        
        # Add Target Complete Date as empty column (always blank for user input)
        ppt_data['target_date'] = ""
        
        # Add status categorization for counting (only if status is mapped)
        if column_mapping.get('status') and column_mapping['status'] in data.columns:
            ppt_data['_status_category'] = ppt_data['status'].apply(categorize_status)
        else:
            # Default to 'non-closed' if no status mapping
            ppt_data['_status_category'] = 'non-closed'
        
        return ppt_data
    
    def validate_required_columns(self, df: pd.DataFrame, column_mapping: Dict[str, str]) -> List[str]:
        """Validate that all required columns are mapped and contain data"""
        errors = []
        required_columns = ['jira', 'description', 'status', 'ranking']
        
        for req_col in required_columns:
            mapped_col = column_mapping.get(req_col)
            if not mapped_col:
                errors.append(f"Required column '{req_col}' is not mapped")
            elif mapped_col not in df.columns:
                errors.append(f"Mapped column '{mapped_col}' not found in data")
            elif df[mapped_col].isna().all():
                errors.append(f"Column '{mapped_col}' contains no data")
        
        return errors